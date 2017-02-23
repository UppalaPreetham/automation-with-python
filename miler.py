#!/usr/bin/python
# -*- coding: utf-8 -*-
import json
import xml.dom.minidom
import os
import glob
import zipfile
import sys
import getopt
import os.path
import shutil
import re
import openpyxl


# Code for decompiling the jar files that are downloaded from the maven through pomn dependencies start

def unzipjar(jar):
    file = zipfile.ZipFile(jar)
    file.extractall(os.path.splitext(jar)[0])


def q(s):
    quote = '"'
    s = quote + s + quote
    return s


def scandirs(path, jadFilePath):
    space = ' '
    opts = '-o -s java -d '
    for currentFile in glob.glob(os.path.join(path, '*')):
        if os.path.isdir(currentFile):
            scandirs(currentFile, jadFilePath)
        elif os.path.splitext(currentFile)[1] == '.class':
            command = q(jadFilePath + "\JAD.exe") + space + opts \
                + q(os.path.dirname(currentFile)) + space \
                + q(currentFile)
            os.system(q(command))
            os.system('del ' + q(currentFile))


def lookforjars(path, jadFilePath):
    for currentFile in glob.glob(os.path.join(path, '*')):
        if os.path.isdir(currentFile):
            (currentFile, jadFilePath)
        elif os.path.splitext(currentFile)[1] == '.jar':
            unzipjar(currentFile)
            scandirs(os.path.splitext(currentFile)[0], jadFilePath)


# Code for decompiling the jar files that are downloaded from the maven through pomn dependencies end

# Class object which holds the details that are parsed from the json file provided

class CompObject:

    art = ''
    groupId = ''
    version = ''

    def __init__(
        self,
        art,
        groupId,
        version,
        ):
        self.art = art
        self.groupId = groupId
        self.version = version
        


# Class object which holds the details after parsing along with count of lines

class CompObjectWithCounts:

    art = ''
    groupId = ''
    version = ''

    def __init__(
        self,
        art,
        groupId,
        version,
        countOfComments,
        countOfBlank,
        countOfCode
        ):
        self.art = art
        self.groupId = groupId
        self.version = version
        self.countOfComments = countOfComments
        self.countOfBlank = countOfBlank
        self.countOfCode = countOfCode


# function which returns the array of comp objects parsed from the json file

def getComponentObjectsList(filepath):
    componentobjectslist = []
    var = 10  # need to be removed for production only for test purpose limiting to one jar
    with open(filepath) as json_data:
        d = json.load(json_data)
    for component in d:
        if component['displayName'] is not None:
            group = ''
            artifact = ''
            version = ''
            for part in component['displayName']['parts']:
                if 'field' in part:
                    if part['field'] == 'Artifact':
                        group = part['value']
                    if part['field'] == 'Group':
                        artifact = part['value']
                    if part['field'] == 'Version':
                        version = part['value']
            var = var - 1  # need to be removed for production only for test purpose limiting to one jar
            if var == 5:  # need to be removed for production only for test purpose limiting to one jar
                break  # need to be removed for production only for test purpose limiting to one jar
            compObj = CompObject(artifact, group, version)
            componentobjectslist.append(compObj)
    return componentobjectslist


# function which builds the pom.xml file

def buildCustomPom(cm):
    
    pom_template_text = \
        """<project xmlns="http://maven.apache.org/POM/4.0.0" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xsi:schemaLocation="http://maven.apache.org/POM/4.0.0 http://maven.apache.org/xsd/maven-4.0.0.xsd">
    <modelVersion>4.0.0</modelVersion>
    <groupId>TestMaven</groupId>
    <artifactId>TestMaven</artifactId>
    <version>0.0.1-SNAPSHOT</version>
    <dependencies>
"""

    
    pom_template_text = pom_template_text + '\t\t<dependency>\n'
    pom_template_text = pom_template_text + '\t\t\t<groupId>'+ cm.groupId + '</groupId>\n'
    pom_template_text = pom_template_text + '\t\t\t<artifactId>'+ cm.art + '</artifactId>\n'
    pom_template_text = pom_template_text + '\t\t\t<version>'+ cm.version + '</version>\n'
    pom_template_text = pom_template_text + '\t\t</dependency>\n'

    pom_template_text = pom_template_text \
        + """   </dependencies>
                    <build>
                        <sourceDirectory>src</sourceDirectory>
                        <plugins>
                            <plugin>
                                <groupId>org.apache.maven.plugins</groupId>
                                <artifactId>maven-dependency-plugin</artifactId>
                                <version>2.7</version>
                                <executions>
                                    <execution>
                                        <id>default-cli</id>
                                        <configuration>
                                            <artifactItems>
                                                <artifactItem>
                                                    <!-- hardcode values, or use properties, depending on what you want
                                                        to do -->
                                                    <groupId>TestMaven</groupId>
                                                    <artifactId>TestMaven</artifactId>
                                                    <version>0.0.1-SNAPSHOT</version>
                                                    <type>[ packaging ]</type>
                                                    <outputDirectory>C:/Hari_2014119/pythonsamples/lib</outputDirectory>
                                                </artifactItem>
                                            </artifactItems>
                                            <!-- other configurations here -->
                                        </configuration>
                                    </execution>
                                </executions>
                            </plugin>
                        </plugins>
                    </build>
                </project>
                """

    text_file = open('pom.xml', 'w')
    text_file.write(pom_template_text)
    text_file.close()


if __name__ == '__main__':

    if len(sys.argv) == 4:
        jsonFilePath = str(sys.argv[1]) # argument which tells the path of the json file
        jadFilePath = str(sys.argv[2])  # argument which tells the path where JAD.exe is present
        libraryPath = str(sys.argv[3])  # argument which tells the path to which jar files has to be copied

        if os.path.exists(jsonFilePath):
            if os.path.exists(jadFilePath + "\JAD.exe"):
                linesToPrint = "Name of the aritifact ******************************* Total no of lines\n";
                componentobjectslist2 = getComponentObjectsList(jsonFilePath)
                compObjectWithLines2= []
                for cm in componentobjectslist2:
                        buildCustomPom(cm)  # call to create the pom.xml in the same directory of python
                        os.system('mvn dependency:copy-dependencies -DoutputDirectory='
                                       + libraryPath)  # call to download all the maven dependencies
                        
                        lookforjars(libraryPath, jadFilePath)  # call to decompile all the jars that are downloaded

                                        # code for calucalting the no of lines in each java file starts
                        
                        totalcommentCount = 0
                        totalblankCount = 0
                        totalcodeCount = 0
                        
                        os.system('cloc.exe --report-file='+cm.art+'-'+cm.version+'.txt '+libraryPath)
                        if os.path.exists(cm.art+'-'+cm.version+'.txt'):
                            for line in open(cm.art+'-'+cm.version+'.txt '):
                                if line.startswith('Java'):
                                    wordList = re.sub("[^\w]", " ",  line).split()
                                    totalblankCount = wordList[2]
                                    totalcommentCount = wordList[3]
                                    totalcodeCount = wordList[4]
                            print('****************************************************************************************************************')
                            print(cm.art,cm.groupId,cm.version,totalcommentCount,totalblankCount,totalcodeCount)
                            print('****************************************************************************************************************')
                            compObjWithLines = CompObjectWithCounts(cm.art,cm.groupId,cm.version,totalcommentCount,totalblankCount,totalcodeCount)
                            compObjectWithLines2.append(compObjWithLines)
                            for dir in os.listdir(libraryPath):
                                    shutil.rmtree(os.path.join(libraryPath,dir),ignore_errors=True)
                            filelist = [ f for f in os.listdir(libraryPath) if f.endswith(".jar") ]
                            for f in filelist:
                                os.remove(libraryPath+"\\"+f)
                                
    
                xfile = openpyxl.load_workbook('OS_Consumption_Metrics.xlsx')

                sheet = xfile.get_sheet_by_name('OS COCOMO')
                i = 5
                for cloc in compObjectWithLines2:
                    sheet.cell(row=i,column=3).value = cloc.groupId
                    sheet.cell(row=i,column=4).value = cloc.art
                    sheet.cell(row=i,column=5).value = cloc.version
                    sheet.cell(row=i,column=7).value = cloc.countOfCode
                    #sheet['E6'].value = cloc.art
                    #sheet['F6'].value = cloc.version
                    #sheet['H6'].value = cloc.countOfCode
                    i = i+1
                    print(cloc.countOfComments,cloc.countOfBlank,cloc.countOfCode)
                xfile.save('OS_Consumption_Metrics.xlsx')
            else:
                print ('JAD file does not exists at the path u specified')
        else:
            print ('JSON file to create the pom.xml is not present ar the path you specified')
    else:

        print ("""Arguments has to be passed in the below order
                      1. JSON file path
                      2. Directory in which jad.exe is present
                      3. Path of the desktop where downloaded jars should be stored""")
